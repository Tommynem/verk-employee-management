#!/usr/bin/env python3
"""Convert Excel time tracking file to CSV format for import.

This script converts the VaWW Excel time tracking format to the CSV format
expected by the Verk Bookkeeping import system.

Usage:
    python convert_excel_to_csv.py input.xlsx output.csv
"""

import argparse
import csv
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


def parse_break_minutes(value) -> int:
    """Convert break value to minutes.

    Args:
        value: timedelta, int (seconds), or None

    Returns:
        Break duration in minutes
    """
    if value is None:
        return 0
    if isinstance(value, timedelta):
        return int(value.total_seconds() / 60)
    if isinstance(value, (int, float)):
        # Assume seconds if numeric
        return int(value / 60)
    return 0


def determine_absence_type(notes_value: str | None, is_vacation: bool) -> tuple[str, str | None]:
    """Determine absence type and remaining notes from Excel values.

    Args:
        notes_value: Value from the "Abwesenheit / Bemerkung" column
        is_vacation: Value from the "Urlaub?" column

    Returns:
        Tuple of (absence_type in German, remaining_notes or None)
    """
    notes = notes_value.strip() if notes_value else ""
    notes_lower = notes.lower()

    # Check vacation flag first
    if is_vacation:
        return "Urlaub", None

    # Check for keywords in notes
    if "wochenende" in notes_lower:
        return "SKIP", None  # Special marker to skip weekends

    if "feiertag" in notes_lower:
        return "Feiertag", None

    if "krank" in notes_lower:
        # Handle "Zeitausgleich - krank" and similar
        return "Krank", None

    if "zeitausgleich" in notes_lower or "gleitzeit" in notes_lower:
        return "Gleitzeit", None

    # Regular work day with optional notes
    return "Keine", notes if notes else None


def format_time(value) -> str:
    """Format time value to HH:MM string.

    Args:
        value: datetime.time, datetime, or None

    Returns:
        Time string in HH:MM format, or empty string
    """
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    return ""


def format_date(value) -> str:
    """Format date value to YYYY-MM-DD string.

    Args:
        value: datetime or date object

    Returns:
        Date string in YYYY-MM-DD format
    """
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return ""


def process_sheet(ws, rows: list, stats: dict):
    """Process a single worksheet and append entries to rows.

    Args:
        ws: openpyxl Worksheet object
        rows: List to append row dictionaries to
        stats: Dictionary to track statistics
    """
    sheet_name = ws.title
    data_start_row = 9  # Data starts at row 9 based on file structure

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        # Column indices (0-based): A=0, B=1, D=3, F=5, O=14, R=17
        date_val = row[0] if len(row) > 0 else None
        start_time = row[1] if len(row) > 1 else None
        end_time = row[3] if len(row) > 3 else None
        break_val = row[5] if len(row) > 5 else None
        notes_val = row[14] if len(row) > 14 else None
        vacation_val = row[17] if len(row) > 17 else False

        # Skip rows without a date
        if date_val is None or not hasattr(date_val, "strftime"):
            continue

        # Parse values
        work_date = format_date(date_val)
        start_str = format_time(start_time)
        end_str = format_time(end_time)
        break_minutes = parse_break_minutes(break_val)

        # Determine absence type and notes
        absence_type, remaining_notes = determine_absence_type(
            str(notes_val) if notes_val else None,
            bool(vacation_val)
        )

        # Skip weekends
        if absence_type == "SKIP":
            stats["skipped_weekends"] += 1
            continue

        # Create row entry
        entry = {
            "Datum": work_date,
            "Startzeit": start_str,
            "Endzeit": end_str,
            "Pause (Min)": str(break_minutes),
            "Abwesenheit": absence_type,
            "Notizen": remaining_notes or "",
        }

        rows.append(entry)
        stats["processed"] += 1

        # Track absence types for summary
        if absence_type != "Keine":
            stats["by_absence_type"][absence_type] = stats["by_absence_type"].get(absence_type, 0) + 1


def convert_excel_to_csv(input_path: Path, output_path: Path) -> dict:
    """Convert Excel file to CSV format.

    Args:
        input_path: Path to input Excel file
        output_path: Path to output CSV file

    Returns:
        Statistics dictionary
    """
    # Load workbook
    wb = openpyxl.load_workbook(input_path, data_only=True)

    # Collect all rows from all sheets
    rows = []
    stats = {
        "sheets": [],
        "processed": 0,
        "skipped_weekends": 0,
        "by_absence_type": {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        before_count = len(rows)
        process_sheet(ws, rows, stats)
        sheet_count = len(rows) - before_count
        stats["sheets"].append({"name": sheet_name, "entries": sheet_count})
        print(f"  Processed sheet '{sheet_name}': {sheet_count} entries")

    # Sort by date
    rows.sort(key=lambda r: r["Datum"])

    # Write CSV with UTF-8 BOM (German Excel compatible)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Datum", "Startzeit", "Endzeit", "Pause (Min)", "Abwesenheit", "Notizen"],
            delimiter=";",
        )
        writer.writeheader()
        writer.writerows(rows)

    return stats


def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel time tracking file to CSV format for import."
    )
    parser.add_argument(
        "input",
        type=Path,
        help="Input Excel file (.xlsx)",
    )
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        default=None,
        help="Output CSV file (default: same name with .csv extension)",
    )

    args = parser.parse_args()

    # Determine output path
    if args.output is None:
        args.output = args.input.with_suffix(".csv")

    # Validate input
    if not args.input.exists():
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    print(f"Converting: {args.input}")
    print(f"Output: {args.output}")
    print()

    # Convert
    stats = convert_excel_to_csv(args.input, args.output)

    # Print summary
    print()
    print("=" * 50)
    print("Conversion Summary")
    print("=" * 50)
    print(f"Total entries: {stats['processed']}")
    print(f"Skipped weekends: {stats['skipped_weekends']}")
    print()
    if stats["by_absence_type"]:
        print("Absence types:")
        for atype, count in sorted(stats["by_absence_type"].items()):
            print(f"  {atype}: {count}")
    print()
    print(f"Output written to: {args.output}")


if __name__ == "__main__":
    main()
